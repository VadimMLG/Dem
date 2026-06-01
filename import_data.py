# -*- coding: utf-8 -*-
"""
СКРИПТ ИМПОРТА ДАННЫХ ДЛЯ ДЕМОНСТРАЦИОННОГО ЭКЗАМЕНА
Автоматически читает Excel-файлы и загружает в MySQL базу данных.

ИНСТРУКЦИЯ:
1. Убедись что XAMPP запущен (MySQL горит зелёным)
2. Импортируй shoe2026_pu.sql в phpMyAdmin (создай базу exam_v5 и импортируй)
3. Укажи ПРАВИЛЬНЫЙ ПУТЬ к папке import ниже (DB_NAME и IMPORT_FOLDER)
4. Запусти этот скрипт: python import_data.py
"""

import os
import sys

# ============================================================
# НАСТРОЙКИ - ИЗМЕНИ ЭТИ 2 СТРОКИ ПОД СВОЙ ВАРИАНТ!
# ============================================================
DB_NAME = "exam_v5"           # <- имя базы данных в phpMyAdmin
IMPORT_FOLDER = r"C:\Users\Вадим\Desktop\DemTest\jackpot\Демонстрационный экзамен\1 июня\415\01.06-02.06\1\Модуль 1\import"
# ============================================================

DB_HOST = "localhost"
DB_USER = "root"
DB_PASS = ""

def install_deps():
    """Установка нужных пакетов если их нет"""
    try:
        import openpyxl
        import mysql.connector
    except ImportError:
        print("Устанавливаю необходимые пакеты...")
        os.system(f"{sys.executable} -m pip install openpyxl mysql-connector-python -q")

install_deps()

import openpyxl
import mysql.connector

def connect():
    return mysql.connector.connect(
        host=DB_HOST, user=DB_USER, password=DB_PASS, database=DB_NAME, charset='utf8mb4'
    )

def read_xlsx(path):
    """Читает xlsx и возвращает список строк (без заголовка)"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # пропускаем заголовок
        if any(cell is not None for cell in row):
            rows.append([str(v).strip() if v is not None else '' for v in row])
    return rows

def safe(val):
    """Безопасное значение для SQL"""
    if val is None or str(val).strip() == '' or str(val).strip() == 'None':
        return None
    return str(val).strip()

def import_users(cursor):
    path = os.path.join(IMPORT_FOLDER, "user_import.xlsx")
    if not os.path.exists(path):
        print(f"  ФАЙЛ НЕ НАЙДЕН: {path}")
        return
    rows = read_xlsx(path)
    cursor.execute("DELETE FROM users_import_raw")
    sql = "INSERT INTO users_import_raw (role_name, full_name, login_text, password_text) VALUES (%s,%s,%s,%s)"
    for row in rows:
        while len(row) < 4:
            row.append('')
        cursor.execute(sql, (safe(row[0]), safe(row[1]), safe(row[2]), safe(row[3])))
    print(f"  users_import_raw: загружено {len(rows)} строк")

def import_products(cursor):
    path = os.path.join(IMPORT_FOLDER, "Tovar.xlsx")
    if not os.path.exists(path):
        print(f"  ФАЙЛ НЕ НАЙДЕН: {path}")
        return
    rows = read_xlsx(path)
    cursor.execute("DELETE FROM products_import_raw")
    sql = """INSERT INTO products_import_raw 
             (article_text, name_text, unit_text, price_text, supplier_text, 
              manufacturer_text, category_text, discount_text, stock_text, description_text, photo_text)
             VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
    for row in rows:
        while len(row) < 11:
            row.append('')
        cursor.execute(sql, tuple(safe(v) for v in row[:11]))
    print(f"  products_import_raw: загружено {len(rows)} строк")

def import_orders(cursor):
    path = os.path.join(IMPORT_FOLDER, "Заказ_import.xlsx")
    if not os.path.exists(path):
        print(f"  ФАЙЛ НЕ НАЙДЕН: {path}")
        return
    rows = read_xlsx(path)
    cursor.execute("DELETE FROM orders_import_raw")
    sql = """INSERT INTO orders_import_raw
             (order_number_text, articles_text, order_date_text, delivery_date_text,
              pickup_point_text, client_fio_text, pickup_code_text, status_text)
             VALUES (%s,%s,%s,%s,%s,%s,%s,%s)"""
    for row in rows:
        while len(row) < 8:
            row.append('')
        cursor.execute(sql, tuple(safe(v) for v in row[:8]))
    print(f"  orders_import_raw: загружено {len(rows)} строк")

def import_pickup_points(cursor):
    path = os.path.join(IMPORT_FOLDER, "Пункты выдачи_import.xlsx")
    if not os.path.exists(path):
        print(f"  ФАЙЛ НЕ НАЙДЕН: {path}")
        return
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            val = str(row[0]).strip() if row[0] is not None else ''
            if val and val != 'None':
                rows.append(val)
    cursor.execute("DELETE FROM pickup_points_import_raw")
    sql = "INSERT INTO pickup_points_import_raw (address_text) VALUES (%s)"
    for addr in rows:
        cursor.execute(sql, (addr,))
    print(f"  pickup_points_import_raw: загружено {len(rows)} строк")

def run_production_script(cursor, conn):
    """Запускает SQL скрипт переноса из raw в боевые таблицы"""
    # Читаем скрипт из папки рядом
    script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SQL_script_raw_to_production.sql")
    if not os.path.exists(script_path):
        print(f"  СКРИПТ НЕ НАЙДЕН: {script_path}")
        print("  Запусти SQL_script_raw_to_production.sql вручную в phpMyAdmin!")
        return
    
    with open(script_path, 'r', encoding='utf-8') as f:
        sql_content = f.read()
    
    # Разбиваем на отдельные команды
    statements = [s.strip() for s in sql_content.split(';') if s.strip()]
    errors = 0
    for stmt in statements:
        try:
            cursor.execute(stmt)
            conn.commit()
        except Exception as e:
            errors += 1
            print(f"  Предупреждение (не критично): {e}")
    
    if errors == 0:
        print("  Скрипт выполнен успешно!")
    else:
        print(f"  Скрипт выполнен с {errors} предупреждениями (обычно это нормально)")

def main():
    print("=" * 60)
    print("  ИМПОРТ ДАННЫХ ДЛЯ ДЕМОНСТРАЦИОННОГО ЭКЗАМЕНА")
    print("=" * 60)
    print(f"\nБаза данных: {DB_NAME}")
    print(f"Папка с файлами: {IMPORT_FOLDER}\n")
    
    if not os.path.exists(IMPORT_FOLDER):
        print(f"ОШИБКА: Папка не найдена!\n{IMPORT_FOLDER}")
        print("\nИзмени переменную IMPORT_FOLDER в начале скрипта!")
        input("\nНажми Enter для выхода...")
        return
    
    try:
        conn = connect()
        cursor = conn.cursor()
        print("Подключение к MySQL: OK\n")
    except Exception as e:
        print(f"ОШИБКА ПОДКЛЮЧЕНИЯ К MySQL: {e}")
        print("\nПроверь:")
        print("  1. XAMPP запущен (MySQL горит зелёным)")
        print(f"  2. База данных '{DB_NAME}' создана в phpMyAdmin")
        print(f"  3. Файл shoe2026_pu.sql импортирован в '{DB_NAME}'")
        input("\nНажми Enter для выхода...")
        return
    
    print("Шаг 1: Загружаем данные в raw-таблицы...")
    try:
        import_users(cursor)
        conn.commit()
    except Exception as e:
        print(f"  Ошибка users: {e}")
    
    try:
        import_products(cursor)
        conn.commit()
    except Exception as e:
        print(f"  Ошибка products: {e}")
    
    try:
        import_orders(cursor)
        conn.commit()
    except Exception as e:
        print(f"  Ошибка orders: {e}")
    
    try:
        import_pickup_points(cursor)
        conn.commit()
    except Exception as e:
        print(f"  Ошибка pickup_points: {e}")
    
    print("\nШаг 2: Переносим данные в боевые таблицы...")
    try:
        run_production_script(cursor, conn)
    except Exception as e:
        print(f"  Ошибка: {e}")
    
    cursor.close()
    conn.close()
    
    print("\n" + "=" * 60)
    print("  ГОТОВО! Теперь запускай проект в Visual Studio (F5)")
    print("=" * 60)
    input("\nНажми Enter для выхода...")

if __name__ == "__main__":
    main()
